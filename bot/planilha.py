"""
planilha.py — Leitura e iteração sobre filtro_alunos.xlsx
"""
from dataclasses import dataclass
from pathlib import Path
from typing import Generator

from loguru import logger
from openpyxl import load_workbook

from bot.config import Config


@dataclass
class Aluno:
    telefone: str
    nome: str
    matricula: str
    linha: int  # linha original na planilha (útil para debug)


def _col_para_indice(letra: str) -> int:
    """Converte letra de coluna (A, B, C...) para índice 1-based."""
    letra = letra.strip().upper()
    resultado = 0
    for c in letra:
        resultado = resultado * 26 + (ord(c) - ord("A") + 1)
    return resultado


def carregar_alunos() -> Generator[Aluno, None, None]:
    """
    Lê a planilha e retorna um gerador de Aluno.
    Ignora linhas com telefone ou nome em branco.
    """
    caminho = Path(Config.PLANILHA_PATH)
    if not caminho.exists():
        raise FileNotFoundError(
            f"Planilha não encontrada: {caminho.resolve()}\n"
            f"Coloque o arquivo filtro_alunos.xlsx na pasta data/"
        )

    col_tel       = _col_para_indice(Config.COLUNA_TELEFONE)
    col_nome      = _col_para_indice(Config.COLUNA_NOME)
    col_matricula = _col_para_indice(Config.COLUNA_MATRICULA)

    wb = load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active

    total = 0
    ignorados = 0

    for idx, row in enumerate(ws.iter_rows(min_row=Config.LINHA_INICIO, values_only=True), start=Config.LINHA_INICIO):
        telefone  = str(row[col_tel - 1]).strip()  if row[col_tel - 1]  is not None else ""
        nome      = str(row[col_nome - 1]).strip() if row[col_nome - 1] is not None else ""
        # Guard: linha pode ter menos colunas que o esperado
        idx_mat   = col_matricula - 1
        matricula = str(row[idx_mat]).strip() if (len(row) > idx_mat and row[idx_mat] is not None) else ""
        # remove decimal do openpyxl (ex: "123456.0" → "123456")
        if matricula.endswith(".0"):
            matricula = matricula[:-2]
        # "None" como string → vazio
        if matricula == "None":
            matricula = ""

        if not telefone or not nome or telefone == "None":
            ignorados += 1
            continue

        # Normaliza telefone: remove espaços, traços, parênteses
        telefone = (
            telefone
            .replace(" ", "")
            .replace("-", "")
            .replace("(", "")
            .replace(")", "")
        )



        total += 1
        yield Aluno(telefone=telefone, nome=nome, matricula=matricula, linha=idx)

    wb.close()
    logger.info(f"Planilha carregada — {total} alunos válidos, {ignorados} linhas ignoradas")
